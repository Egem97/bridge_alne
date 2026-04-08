import io
import json
from datetime import datetime

import openpyxl
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from asgiref.sync import sync_to_async

from .utils import NetSuiteClient
from .models import UploadHistory
from .services.pipeline import process_upload
from .services.payload import reorder_payload
from .services.master_data import MasterDataLoader
from dashboard.decorators import role_required

# Which planilla types each role can access
ROLE_PLANILLA_MAP = {
    'Nomina Empleado': ['empleados'],
    'Nomina Obreros':  ['obreros'],
    'Owner':           ['empleados', 'obreros'],
}


def _get_allowed_types(user_role):
    return ROLE_PLANILLA_MAP.get(user_role, [])


@login_required
@role_required(allowed_roles=['Nomina Empleado', 'Nomina Obreros'])
async def upload_excel_view(request, planilla_type=None):

    def get_user_role():
        try:
            return request.user.profile.role.description
        except Exception:
            return None

    user_role = await sync_to_async(get_user_role)()
    allowed_types = _get_allowed_types(user_role)

    # Default planilla_type: first allowed type for the user
    if planilla_type is None or planilla_type not in allowed_types:
        planilla_type = allowed_types[0] if allowed_types else 'empleados'

    if request.method == 'POST':

        # ── Confirm flow ────────────────────────────────────────────────────
        if request.content_type == 'application/json':
            try:
                body = json.loads(request.body)
                if body.get('action') == 'confirm_upload':
                    upload_id = body.get('upload_id')

                    def run_confirm():
                        try:
                            history = UploadHistory.objects.get(id=upload_id)
                        except UploadHistory.DoesNotExist:
                            return {"valid": False, "error": "Registro de carga no encontrado"}

                        # Reorder keys: PostgreSQL JSONB does not preserve insertion order
                        payload = reorder_payload(history.netsuite_payload)

                        history.status = UploadHistory.Status.SENDING
                        history.save(update_fields=['status', 'modified'])

                        client = NetSuiteClient()
                        try:
                            response = client.restlet(payload)
                            history.netsuite_response = response

                            body_items = (
                                response if isinstance(response, list)
                                else response.get('body', []) if isinstance(response, dict)
                                else []
                            )
                            ns_errors = [
                                item.get('msg', 'Error desconocido')
                                for item in body_items
                                if isinstance(item, dict) and item.get('status') == 'error'
                            ]

                            if ns_errors:
                                history.status = UploadHistory.Status.ERROR
                                history.error_message = '; '.join(ns_errors)
                                history.save()
                                return {
                                    "valid": False,
                                    "error": '; '.join(ns_errors),
                                    "ns_results": [{"status": "netsuite_error", "error": msg} for msg in ns_errors],
                                }

                            history.status = UploadHistory.Status.SUCCESS
                            history.save()
                            return {
                                "valid": True,
                                "ns_results": [{"status": "netsuite_response", "data": response}],
                                "row_count": history.row_count,
                            }
                        except Exception as ns_e:
                            history.status = UploadHistory.Status.ERROR
                            history.error_message = str(ns_e)
                            history.save()
                            return {
                                "valid": False,
                                "error": str(ns_e),
                                "ns_results": [{"status": "netsuite_error", "error": str(ns_e)}],
                            }

                    result = await sync_to_async(run_confirm)()
                    return JsonResponse(result)

            except Exception as e:
                return JsonResponse(
                    {"valid": False, "error": f"Error procesando solicitud: {str(e)}"},
                    status=400
                )

        # ── Upload flow ─────────────────────────────────────────────────────
        elif request.FILES.get('file'):
            uploaded_file = request.FILES['file']
            file_content = uploaded_file.read()
            upload_planilla_type = request.POST.get('planilla_type', planilla_type)

            # Enforce role-based planilla restriction
            if upload_planilla_type not in allowed_types:
                return JsonResponse(
                    {"valid": False, "errors": [{"level": "error", "category": "auth",
                     "message": f"No tienes permiso para cargar planillas de tipo '{upload_planilla_type}'.",
                     "details": {}}]},
                    status=403
                )

            def process():
                return process_upload(file_content, upload_planilla_type)

            result = await sync_to_async(process)()

            if result.get('valid'):
                transformed_data = result.pop('_transformed_data', None)
                transformed_columns = result.pop('_transformed_columns', None)

                def save_history():
                    return UploadHistory.objects.create(
                        user=request.user,
                        file_name=uploaded_file.name,
                        file_size=uploaded_file.size,
                        planilla_type=upload_planilla_type,
                        subsidiary_name=result.get('subsidiary_name'),
                        status=UploadHistory.Status.PREVIEW,
                        row_count=result.get('row_count', 0),
                        total_debit=result.get('total_debit'),
                        total_credit=result.get('total_credit'),
                        netsuite_payload=result.get('netsuite_payload'),
                        transformed_data=transformed_data,
                        transformed_columns=transformed_columns,
                    )

                history = await sync_to_async(save_history)()
                result['upload_id'] = str(history.id)
            else:
                transformed_data = result.pop('_transformed_data', None)
                transformed_columns = result.pop('_transformed_columns', None)

                def save_error_history():
                    return UploadHistory.objects.create(
                        user=request.user,
                        file_name=uploaded_file.name,
                        file_size=uploaded_file.size,
                        planilla_type=upload_planilla_type,
                        subsidiary_name=result.get('subsidiary_name'),
                        status=UploadHistory.Status.VALIDATION_ERROR,
                        validation_errors=result.get('errors', []),
                        error_message=str(result.get('errors', [])),
                        transformed_data=transformed_data,
                        transformed_columns=transformed_columns,
                    )

                try:
                    history = await sync_to_async(save_error_history)()
                    # If we have transformed data, expose upload_id so user can download
                    if transformed_data:
                        result['upload_id'] = str(history.id)
                except Exception as db_err:
                    import logging
                    logging.getLogger(__name__).error(f"Error saving error history: {db_err}")

            return JsonResponse(result)

    return await sync_to_async(render)(
        request,
        'yw_oracle/upload.html',
        {
            'planilla_type': planilla_type,
            'allowed_types': allowed_types,
        }
    )


@login_required
@role_required(allowed_roles=['Nomina Empleado', 'Nomina Obreros'])
async def download_excel_view(request, upload_id):
    """Returns the full transformed DataFrame as an Excel file."""

    def get_user_role():
        try:
            return request.user.profile.role.description
        except Exception:
            return None

    user_role = await sync_to_async(get_user_role)()
    allowed_types = _get_allowed_types(user_role)

    def build_excel():
        try:
            history = UploadHistory.objects.get(id=upload_id)
        except UploadHistory.DoesNotExist:
            return None, "Registro no encontrado"

        if history.planilla_type not in allowed_types:
            return None, "Sin permiso para descargar este archivo"

        rows = history.transformed_data or []
        cols = history.transformed_columns or []

        if not rows or not cols:
            return None, "No hay datos transformados disponibles"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transformacion"

        # Header row with bold style
        header_font = openpyxl.styles.Font(bold=True)
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(cols, start=1):
                value = row.get(col_name)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"transformacion_{history.file_name.rsplit('.', 1)[0]}_{history.planilla_type}.xlsx"
        return buffer, filename

    result, meta = await sync_to_async(build_excel)()

    if result is None:
        return JsonResponse({"error": meta}, status=404)

    response = HttpResponse(
        result.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{meta}"'
    return response


@login_required
@role_required(allowed_roles=['Nomina Empleado', 'Nomina Obreros'])
async def netsuite_query_view(request):
    query = request.GET.get('q', "SELECT * FROM Transaction WHERE Rownum <= 5")

    def run_query():
        client = NetSuiteClient()
        return client.execute_suiteql(query)

    try:
        data = await sync_to_async(run_query)()
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def _build_enrichment_maps():
    """
    Build reverse lookup dicts from local Excel/Parquet files so the detail view
    can resolve NetSuite internal IDs to human-readable names without extra API calls.
    All failures are swallowed so a missing file never breaks the response.
    """
    maps = {
        'subsidiary': {},   # id (int) → name (str)
        'ceco':       {},   # id (int) → name (str)
        'account':    {},   # id (int) → {'code': str, 'name': str}
        'actividad':  {},   # id (int) → name (str)
    }

    try:
        sub_dict = MasterDataLoader.get_subsidiary_dict()   # {name: id}
        maps['subsidiary'] = {int(v): k for k, v in sub_dict.items()}
    except Exception:
        pass

    try:
        ceco_map = MasterDataLoader.get_ceco_map()           # {sub_id: {name: ceco_id}}
        for cecos in ceco_map.values():
            for name, ceco_id in cecos.items():
                try:
                    maps['ceco'][int(ceco_id)] = name
                except (ValueError, TypeError):
                    pass
    except Exception:
        pass

    try:
        accounts_df = MasterDataLoader.get_accounts_excel()
        for _, row in accounts_df.iterrows():
            try:
                maps['account'][int(row['id_cuenta'])] = {
                    'code': str(row.get('CUENTA CONTABLE', '') or ''),#NAME CUENTA
                    'name': str(row.get('DESCRIPCION', '') or ''),
                }
            except (ValueError, TypeError):
                pass
    except Exception:
        pass

    try:
        act_df = MasterDataLoader.get_actividad_table()
        if 'id_actividad' in act_df.columns:
            for _, row in act_df.iterrows():
                try:
                    maps['actividad'][int(row['id_actividad'])] = str(row.get('Actividad del Proyecto', '') or '')
                except (ValueError, TypeError):
                    pass
    except Exception:
        pass

    return maps


@login_required
async def transactions_view(request):

    def get_subsidiaries():
        sub_dict = MasterDataLoader.get_subsidiary_dict()
        return sorted(
            [{"name": name, "id": sid} for name, sid in sub_dict.items()],
            key=lambda x: x["name"]
        )

    if request.method == 'GET':
        subsidiaries = await sync_to_async(get_subsidiaries)()
        return await sync_to_async(render)(
            request,
            'yw_oracle/transactions.html',
            {'subsidiaries': subsidiaries}
        )

    if request.method == 'POST' and request.content_type == 'application/json':
        try:
            body = json.loads(request.body)
            action = body.get('action')

            if action == 'list':
                date_from = body.get('date_from', '')
                
                date_to = body.get('date_to', '')
                
                subsidiary_id = body.get('subsidiary_id')
                
                try:
                    datetime.strptime(date_from, '%Y-%m-%d')
                    datetime.strptime(date_to, '%Y-%m-%d')
                except (ValueError, TypeError):
                    return JsonResponse({'error': 'Fechas inválidas. Use el formato YYYY-MM-DD.'}, status=400)

                try:
                    subsidiary_id = int(subsidiary_id)
                except (ValueError, TypeError):
                    return JsonResponse({'error': 'Subsidiaria inválida.'}, status=400)

                query = (
                    "SELECT "
                    "t.id AS transaction_id, "
                    "t.tranid AS numero_asiento, "
                    "t.trandate AS fecha, "
                    "t.postingperiod AS periodo, "
                    "t.memo AS descripcion, "
                    "t.createddate AS fecha_creacion, "
                    "t.lastmodifieddate AS ultima_modificacion "
                    "FROM transaction t "
                    "WHERE t.type = 'Journal' "
                    "AND t.voided = 'F' "
                    f"AND TRUNC(t.createddate) >= TO_DATE('{date_from}', 'YYYY-MM-DD') "
                    f"AND TRUNC(t.createddate) <= TO_DATE('{date_to}', 'YYYY-MM-DD') "
                    f"AND t.subsidiary = {subsidiary_id} "
                    "AND t.memo NOT LIKE 'DET%' "
                    "AND t.memo NOT LIKE 'Det%' "
                    "ORDER BY t.createddate DESC, t.id"
                )

                def run_list():
                    client = NetSuiteClient()
                    return client.execute_suiteql(query)

                result = await sync_to_async(run_list)()
                return JsonResponse(result, safe=False)

            elif action == 'detail':
                try:
                    transaction_id = int(body.get('transaction_id', 0))
                except (ValueError, TypeError):
                    return JsonResponse({'error': 'ID de transacción inválido.'}, status=400)

                if transaction_id <= 0:
                    return JsonResponse({'error': 'ID de transacción inválido.'}, status=400)

                query = (
                    "SELECT "
                    "t.id AS transaction_id, "
                    "t.tranid AS numero_asiento, "
                    "t.trandate AS fecha, "
                    "t.postingperiod AS periodo, "
                    "t.memo AS descripcion, "
                    "t.createddate AS fecha_creacion, "
                    "t.lastmodifieddate AS ultima_modificacion, "
                    "tl.id AS linea_id, "
                    "tl.linesequencenumber AS numero_linea, "
                    "tl.account AS id_cuenta, "
                    "CASE WHEN tl.amount > 0 THEN tl.amount ELSE 0 END AS debito, "
                    "CASE WHEN tl.amount < 0 THEN ABS(tl.amount) ELSE 0 END AS credito, "
                    "t.subsidiary AS id_subsidiary, "
                    "tl.department AS id_department, "
                    "tl.cseg_actividad AS id_actividad "
                    "FROM transaction t "
                    "INNER JOIN transactionline tl ON t.id = tl.transaction "
                    "WHERE t.type = 'Journal' "
                    "AND t.voided = 'F' "
                    "AND tl.account IS NOT NULL "
                    f"AND t.id = {transaction_id} "
                    
                    "ORDER BY tl.linesequencenumber"
                )

                def run_detail():
                    def _int(val):
                        try:
                            return int(val)
                        except (TypeError, ValueError):
                            return None

                    client = NetSuiteClient()
                    ns_data = client.execute_suiteql(query)
                    items = ns_data.get('items', [])

                    enrich = _build_enrichment_maps()

                    # Collect account IDs not resolved by local files
                    missing_ids = {
                        _int(item.get('id_cuenta'))
                        for item in items
                        if _int(item.get('id_cuenta')) is not None
                        and _int(item.get('id_cuenta')) not in enrich['account']
                    }

                    # Fetch missing accounts from NetSuite in one batch query
                    if missing_ids:
                        ids_str = ', '.join(str(i) for i in missing_ids)
                        acc_query = (
                            "SELECT id, accountnumber, name "
                            "FROM account "
                            f"WHERE id IN ({ids_str})"
                        )
                        try:
                            acc_data = client.execute_suiteql(acc_query)
                            for acc in acc_data.get('items', []):
                                acc_id = _int(acc.get('id'))
                                if acc_id is not None:
                                    enrich['account'][acc_id] = {
                                        'code': str(acc.get('accountnumber') or ''),
                                        'name': str(acc.get('name') or ''),
                                    }
                        except Exception:
                            pass  # Missing names are non-fatal

                    total_debit = 0.0
                    total_credit = 0.0
                    per_account = {}   # id_cuenta → {code, name, debit, credit}
                    enriched = []

                    for item in items:
                        debit  = float(item.get('debito')  or 0)
                        credit = float(item.get('credito') or 0)
                        total_debit  += debit
                        total_credit += credit

                        id_cuenta     = _int(item.get('id_cuenta'))
                        id_subsidiary = _int(item.get('id_subsidiary'))
                        id_department = _int(item.get('id_department'))
                        id_actividad  = _int(item.get('id_actividad'))

                        acc_info   = enrich['account'].get(id_cuenta, {}) if id_cuenta is not None else {}
                        sub_name   = enrich['subsidiary'].get(id_subsidiary, str(id_subsidiary or ''))
                        dept_name  = enrich['ceco'].get(id_department, str(id_department or ''))
                        act_name   = enrich['actividad'].get(id_actividad, str(id_actividad or ''))

                        # Per-account totals
                        acc_key = id_cuenta or 0
                        if acc_key not in per_account:
                            per_account[acc_key] = {
                                'id': acc_key,
                                'code': acc_info.get('code', str(id_cuenta or '')),
                                'name': acc_info.get('name', ''),
                                'debit': 0.0,
                                'credit': 0.0,
                            }
                        per_account[acc_key]['debit']  += debit
                        per_account[acc_key]['credit'] += credit

                        enriched.append({
                            **item,
                            'debito':              round(debit, 2),
                            'credito':             round(credit, 2),
                            'cuenta_codigo':       acc_info.get('code', str(id_cuenta or '')),
                            'cuenta_nombre':       acc_info.get('name', ''),
                            'subsidiaria_nombre':  sub_name,
                            'departamento_nombre': dept_name,
                            'actividad_nombre':    act_name,
                        })

                    for v in per_account.values():
                        v['debit']  = round(v['debit'],  2)
                        v['credit'] = round(v['credit'], 2)

                    metrics = {
                        'total_debit':  round(total_debit,  2),
                        'total_credit': round(total_credit, 2),
                        'per_account':  sorted(per_account.values(), key=lambda x: x['code']),
                    }

                    return {'items': enriched, 'metrics': metrics, 'count': len(enriched)}

                result = await sync_to_async(run_detail)()
                return JsonResponse(result, safe=False)

            else:
                return JsonResponse({'error': 'Acción no válida.'}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido.'}, status=405)
